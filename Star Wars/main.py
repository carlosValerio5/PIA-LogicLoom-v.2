#!/usr/bin/env python3

import json
import os
import requests
import sys
import time
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
from statistics import mode, mean, median
import shutil
from colorama import Fore, Style, init  # Se usará colorama para cambiar los colores de las letras
from opexl import abrirex

#user defined modules
from modulo_graficas import mostrar_graficoApa, mostrar_graficoEsp, mostrar_graficaDen
from modulo_calculos import moda_planeta, mediana_peso, media_clima
import PauseAndCls  # Es un módulo creado para limpiar o pausar pantalla
import Carpetas_Archivos as CA 
import Menus 
from Data import connectToDb

init()  # Esto inicializa la salida de texto de colores en la terminal



def rutaAH():  # Función para crear la ruta del archivo historial
    direccionA = CA.DireccionJoin("Historial.txt")
    return direccionA


def crearH(urlf, varf):  # Función para crear el historial
    fecha = datetime.now()
    fechaC = str(fecha.year) + "-" + str(fecha.month) + "-" + str(fecha.day)
    horaC = str(fecha.hour) + ":" + str(fecha.minute) + ":" + str(fecha.second)
    with open(rutaAH(), "a") as file: 
        file.write("\n\nSe realizó una consulta el: " + fechaC + " " + horaC)
        file.write("\nBúsqueda: "+ urlf)
        file.write("\nNombre busqueda: " + varf)

# Función para requests a la API
def requ(tipo):
    # Se concatena la extensión para obtener request
    var = input("Ingrese un nombre:")
    url = "https://swapi.dev/api/" + str(tipo) + "/?search=" + var
    try:
        r = requests.get(url)
        crearH(url, var) # Si la consulta ya esta lista se creara el archivos
        # Salir del programa en caso de que no se logre conexión
    except Exception as e1:
        print("Un error ha ocurrido, saliendo")
        print(e1)
        exit()
    return r

# Función para imprimir resultado de request
def show(obj, opB , iP, iPla, iE):
    cambio = False
    info = json.loads(obj.content)
    #si no se obtienen resultados count es 0
    if info["count"] == 0:
        print("No se encontraron coincidencias")
    else:
        print("Resultados: ", info["count"])
        # El contenido que nos interesa está dentro de "results"
        result = info["results"]  # results es una lista de diccionarios
        for i in result:
            for x, y in i.items():
                print(x, ":", y, "\n")
        resp = input("Quieres guardar los datos de la consulta si o no?")
        while resp.upper() != "SI" and resp.upper() != "NO":
            resp = input("Escribe SI o NO: ")
        if resp.upper() == "SI":
            cambio = crearExc(info, opB, iP, iPla, iE)
    return cambio


def carpetas(): #Crea las carpetas donde se guardara los excel
    nombre_Carpetas = ["Personajes", "Planetas", "Especies"]
    for carpeta in nombre_Carpetas:
        ruta = CA.DireccionJoin(carpeta)
        try: 
            shutil.rmtree(ruta)
        except FileNotFoundError:
            print(end="")
        os.makedirs(ruta)


def crearExc(datos, opB, iP, iPla, iE): #Funcion para crear los excel
    cambio = False
    libro = Workbook()
    hoja = libro.active
    numLinea = 1
    hoja.cell(numLinea, 1, "Info Busqueda")
    result = datos["results"] #Recurimos a todos los diccionarios de la request
    for i in result:
        numLinea +=1
        for x,y in i.items():
            alineacion = Alignment(horizontal="left")
            celda = hoja.cell(numLinea,4)
            celda.alignment = alineacion
            hoja.cell(numLinea, 3, x + ":")
            if isinstance(y, list):
                numLinea += 1
                for link in y:
                    hoja.cell(numLinea,4, "- " +  str(link))
                    numLinea +=1
                    continue
            elif str(y).isnumeric():  #? Verificar si y es un número
                hoja.cell(numLinea, 4, int(y))
                numLinea += 1
            else:
                hoja.cell(numLinea, 4, str(y))
                numLinea += 1
            hoja.column_dimensions['C'].auto_size = True
            hoja.column_dimensions['D'].auto_size = True
    if opB == 1:
        ruta = os.path.join(CA.DireccionJoin("Personajes"),"Personajes_"+ str(iP) + ".xlsx")
        libro.save(ruta)
        cambio = True
    elif opB == 2:
        ruta = os.path.join(CA.DireccionJoin("Especies"),"Especies_" + str(iE) + ".xlsx")
        libro.save(ruta)
        cambio = True
    elif opB == 3:
        ruta = os.path.join(CA.DireccionJoin("Planetas"),"Planetas_"+ str(iPla) + ".xlsx")
        libro.save(ruta)
        cambio = True
    return cambio


if __name__ == "__main__":  # AQUÍ IRÁ TODO LO QUE NO SEA FUNCIÓN

    #Begin connection witg database
    connection = connectToDb()

    iP = 1
    iE = 1
    iPla = 1
    carpetas()
    try:  # Líneas del 121 al 126 que borran el archivo historial de la ejecución pasada
        os.remove(rutaAH())
    except FileNotFoundError:
        print(end="")
    except Exception:
        print(end="")
    #Programa desde la terminal linea 144 a 170
    if len(sys.argv) > 2:  #Comprobamos si los argumentos recibidos son los necesario
        try:#Si el programa recibe un parametro como una letra o decimal mandara error
            if int(sys.argv[1]) == 1:
                if int(sys.argv[2]) == 1:
                    mediana_peso()
                elif int(sys.argv[2]) == 2:
                    moda_planeta()
                elif int(sys.argv[2]) == 3:
                    media_clima()
                sys.exit()
            elif int(sys.argv[1]) == 2:
                if int(sys.argv[2]) == 1:
                    mostrar_graficoApa()
                elif int(sys.argv[2]) == 2:
                    mostrar_graficoEsp()
                elif int(sys.argv[2]) == 3:
                    mostrar_graficaDen()
                sys.exit()
        except ValueError:
            print("\nLos argumentos deben ser números entros")
            print("\nArgumento 1 [1-2] \tArgumento 2 [1-3]")
            sys.exit(1)
        except Exception:
            print("Error")
            sys.exit(1)
        finally:
            sys.exit()

    else:
        while True:
            op = Menus.lecturaOp()
            if op == 1:
                Menus.menu2()
                opB = Menus.lecturaOpB()
                if opB == 1:
                    print("Búsqueda por Personaje")
                    resp = requ("people")
                elif opB == 2:
                    print("Búsqueda por Especie")
                    resp = requ("species")
                else:
                    print("Búsqueda por Planeta")
                    resp = requ("planets")
                cambio = show(resp, opB, iP, iPla, iE)
                if cambio:
                    if opB == 1:
                        iP +=1
                    if opB == 2: 
                        iE += 1
                    if opB == 3:
                        iPla += 1
                PauseAndCls.pausa()
            elif op == 2:
                try:
                    with(open(rutaAH(),"r")) as historial:
                        print("Historial", end="")
                        for line in historial:
                            print(line, end="")
                        sys.stdout.flush()
                        print()
                        PauseAndCls.pausa()
                except FileNotFoundError: #? Si entra aqui significa que no hay historial
                    print(Fore.RED + Style.BRIGHT + "Aun no tienes datos con el historial" + Style.RESET_ALL)
                    time.sleep(2)
                else:
                    abrirex()
            elif op == 3:
                #Aqui van los calculos y las graficas
                opCalGraf = Menus.menucalgraf()
                if opCalGraf == 1:
                    opCalcs = Menus.menucalcs()
                    if opCalcs == 1:
                        mediana_peso()
                    elif opCalcs == 2:
                        moda_planeta()
                    else:
                        media_clima()
                    PauseAndCls.pausa()
                else:
                    opGrafs = Menus.menugrafs()
                    if opGrafs == 1:
                        mostrar_graficoApa()
                    elif opGrafs == 2:
                        mostrar_graficoEsp()
                    else:
                        mostrar_graficaDen()
                    PauseAndCls.pausa()
            else:
                print("Saliendo", end="")
                sys.stdout.flush()
                time.sleep(1)
                print(".", end="")
                sys.stdout.flush()
                time.sleep(1)
                print(".", end="")
                sys.stdout.flush()
                time.sleep(1)
                print(".", end="")
                break
