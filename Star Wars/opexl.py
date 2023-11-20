import os
import re

from openpyxl import load_workbook

import PauseAndCls

def abrirex():
  location = os.getcwd()
  var1 = len(os.listdir(os.path.join(location, "Personajes")))
  var2 = len(os.listdir(os.path.join(location, "Especies")))
  var3 = len(os.listdir(os.path.join(location, "Planetas")))
  if var1 != 0 or var2 != 0 or var3 != 0:
      var = input("Desea abrir un archivo en especifico? [y/n]")
      while var.lower() != 'y' and var.lower() != 'n':
          print("Ingrese y o n")
          var = input("Desea abrir un archivo en especifico? [y/n]")
      if var == "y":
          for x, y, z in os.walk(location):
              if x == os.path.join(location, "Personajes"):
                  print("Carpeta Personajes:", "\n")
                  if len(z) == 0:
                      print("Aun no hay archivos en Personajes")
                      continue
                  for i in z:
                      print(i)
              elif x == os.path.join(location, "Especies"):
                  print("Carpeta Especies:", "\n")
                  if len(z) == 0:
                      print("Aun no hay archivos en Especies")
                      continue
                  for i in z:
                      print(i)
              elif x == os.path.join(location, "Planetas"):
                  print("Carpeta Planetas:", "\n")
                  if len(z) == 0:
                      print("Aun no hay archivos en Planetas")
                      continue
                  for i in z:
                      print(i)
          #Uso de expresiones regulares para nombre de carpeta
          carp = input("Ingrese el nombre de una carpeta:")
          reg = r'\S(\D+[a-zA-Z]+\D+)\S'
          carp = " " + "1" + carp + "1" + " "
          exp = re.compile(reg)
          nomcarp = exp.search(carp)
          try:
              carpeta = nomcarp.group(1).strip()
          except AttributeError:
              #En caso de que no haya matches = None
              print("No hay matches")
              carpeta = None
          while carpeta != "Especies" and carpeta != "Planetas" and carpeta != "Personajes" or carpeta is None:
              print("Nombre de carpeta no valido")
              carp = input("Ingrese el nombre de una carpeta:")
              carp = " " + "1" + carp + "1" + " "
              nomcarp = exp.search(carp)
              try:
                  carpeta = nomcarp.group(1).strip()
              except AttributeError:
                  print("No hay matches")
                  carpeta = None
          while True:
              nom = input("Ingrese nombre de archivo: \t")
              if len(nom) <=0:
                  print("El nombre no es valido")
                  continue
              if not ".xlsx" in nom:
                  nom += ".xlsx"
                  break
              break
          dir = os.path.join(location, carpeta, nom)
          try:
              data = load_workbook(dir)
          except FileNotFoundError:
              print("El archivo no existe")
              abrirex()
          else:
              hoja = data.active
              for row in range(0, hoja.max_row):
                  for col in hoja.iter_cols(1, hoja.max_column):
                      if col[row].value == None:
                          continue
                      print(col[row].value, end="\t")
                  print()
              PauseAndCls.pausa()
  else:
      print("No hay archivos disponibles")
    